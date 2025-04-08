import cx_Oracle
import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from tkcalendar import DateEntry  
from PIL import Image, ImageTk

class DatePickerFrame(tk.Frame):
    def __init__(self, master, **kwargs):
        super().__init__(master)
        try:
            image = Image.open("export data.ico")
            self.icon = ImageTk.PhotoImage(image)
        except Exception as e:
            print("Nu s-a găsit fișierul export data.ico:", e)
            self.icon = None
        if self.icon:
            icon_label = tk.Label(self, image=self.icon)
            icon_label.pack(side="left", padx=(0, 5))
        self.date_entry = DateEntry(self, **kwargs)
        self.date_entry.pack(side="left")

    def get(self):
        return self.date_entry.get()

def connect_to_oracle():
    try:
        dsn_tns = cx_Oracle.makedsn('130.61.87.48', '1521', service_name='deigdb24_pdb1.sub11141353020.vcngarade24.oraclevcn.com')
        connection = cx_Oracle.connect(user='gara', password='gara', dsn=dsn_tns)
        return connection
    except cx_Oracle.DatabaseError as e:
        messagebox.showerror("Conexiune DB", f"Eroare: {e}")
        return None

def fetch_data_from_casalenta(connection, start_date, end_date):
    try:
        cursor = connection.cursor()
        query = '''
        SELECT DATA, SUMA_PRIM, PAY_TYPE, CASA
        FROM T1CASALENTA
        WHERE DATA BETWEEN TO_DATE(:start_date, 'YYYY-MM-DD') 
                       AND TO_DATE(:end_date, 'YYYY-MM-DD')
        '''
        cursor.execute(query, start_date=start_date, end_date=end_date)
        data = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        return data, columns
    except cx_Oracle.DatabaseError as e:
        messagebox.showerror("Eroare SQL", f"Error while fetching data: {e}")
        return None, None
    finally:
        cursor.close()

def generate_report(data, columns, report_name):
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Raport Casalenta"
        for col_num, column in enumerate(columns, start=1):
            sheet.cell(row=1, column=col_num, value=column)
        for row_num, row in enumerate(data, start=2):
            for col_num, value in enumerate(row, start=1):
                sheet.cell(row=row_num, column=col_num, value=value)
        workbook.save(report_name)
        messagebox.showinfo("Export", f"Raportul a fost salvat cu succes în {report_name}")
    except Exception as e:
        messagebox.showerror("Export", f"Eroare la generarea raportului: {e}")

class DataExportApp:
    def __init__(self, master, data, columns):
        self.master = master
        self.master.title("Vizualizare și Export Date - T1CASALENTA")
        self.data = data
        self.columns = columns
        self.filtered_data = data  
        self.create_widgets()
        self.populate_treeview(self.filtered_data)

    def create_widgets(self):
        filter_frame = ttk.LabelFrame(self.master, text="Filtrare date")
        filter_frame.pack(fill="x", padx=10, pady=5)
        self.filter_vars = {}
        for idx, col in enumerate(self.columns):
            ttk.Label(filter_frame, text=f"{col}:").grid(row=idx, column=0, padx=5, pady=2, sticky="w")
            if col == "DATA":
                entry = DatePickerFrame(filter_frame, date_pattern='yyyy-mm-dd')
                entry.grid(row=idx, column=1, padx=5, pady=2, sticky="w")
            else:
                entry = ttk.Entry(filter_frame)
                entry.grid(row=idx, column=1, padx=5, pady=2, sticky="w")
            self.filter_vars[col] = entry
        button_frame = ttk.Frame(filter_frame)
        button_frame.grid(row=len(self.columns), column=0, columnspan=2, pady=5)
        filter_button = ttk.Button(button_frame, text="Aplică Filtrare", command=self.apply_filter)
        filter_button.pack(side="left", padx=5)
        clear_filter_button = ttk.Button(button_frame, text="Resetare Filtrare", command=self.reset_filter)
        clear_filter_button.pack(side="left", padx=5)
        self.tree = ttk.Treeview(self.master, columns=self.columns, show="headings")
        for col in self.columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120)
        self.tree.pack(fill="both", expand=True, padx=10, pady=5)
        export_button = ttk.Button(self.master, text="Exportă în Excel", command=self.export_to_excel)
        export_button.pack(pady=5)

    def populate_treeview(self, data):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for row in data:
            self.tree.insert("", "end", values=row)

    def apply_filter(self):
        filtered = self.data
        for idx, col in enumerate(self.columns):
            filter_value = self.filter_vars[col].get().strip()
            if filter_value:
                filtered = [row for row in filtered if filter_value.lower() in str(row[idx]).lower()]
        self.filtered_data = filtered
        self.populate_treeview(self.filtered_data)

    def reset_filter(self):
        for col in self.columns:
            if isinstance(self.filter_vars[col], DatePickerFrame):
                self.filter_vars[col].date_entry.set_date(datetime.now())
            else:
                self.filter_vars[col].delete(0, tk.END)
        self.filtered_data = self.data
        self.populate_treeview(self.filtered_data)

    def export_to_excel(self):
        report_name = "raport_casalenta.xlsx"
        generate_report(self.filtered_data, self.columns, report_name)

def load_data(root, start_date_picker, end_date_picker):
    start_date = start_date_picker.get().strip()
    end_date = end_date_picker.get().strip()
    if not start_date or not end_date:
        messagebox.showwarning("Input", "Te rog să completezi ambele câmpuri pentru perioada de extragere.")
        return
    try:
        datetime.strptime(start_date, '%Y-%m-%d')
        datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        messagebox.showerror("Input", "Formatul datelor trebuie să fie YYYY-MM-DD.")
        return
    connection = connect_to_oracle()
    if connection:
        data, columns = fetch_data_from_casalenta(connection, start_date, end_date)
        connection.close()
        if data is not None and columns is not None:
            if not data:
                messagebox.showinfo("Date", "Nu există date pentru perioada selectată.")
                return
            data_window = tk.Toplevel(root)
            app = DataExportApp(data_window, data, columns)
        else:
            messagebox.showerror("Date", "Nu s-au preluat datele din tabelul T1CASALENTA.")
    else:
        messagebox.showerror("Conexiune", "Nu s-a putut stabili conexiunea la baza de date.")

def main():
    root = tk.Tk()
    root.title("Setare Perioadă de Extragere - T1CASALENTA")
    try:
        root.iconbitmap("export data.ico")
    except Exception as e:
        print("Nu s-a găsit fișierul export data.ico:", e)
    frame = ttk.Frame(root, padding=10)
    frame.pack(fill="x", expand=True)
    ttk.Label(frame, text="Data start (YYYY-MM-DD):").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    start_date_picker = DatePickerFrame(frame, date_pattern='yyyy-mm-dd')
    start_date_picker.grid(row=0, column=1, padx=5, pady=5)
    ttk.Label(frame, text="Data end (YYYY-MM-DD):").grid(row=1, column=0, padx=5, pady=5, sticky="w")
    end_date_picker = DatePickerFrame(frame, date_pattern='yyyy-mm-dd')
    end_date_picker.grid(row=1, column=1, padx=5, pady=5)
    load_button = ttk.Button(frame, text="Extrage Date", command=lambda: load_data(root, start_date_picker, end_date_picker))
    load_button.grid(row=2, column=0, columnspan=2, padx=5, pady=10)
    root.mainloop()

if __name__ == "__main__":
    main()
